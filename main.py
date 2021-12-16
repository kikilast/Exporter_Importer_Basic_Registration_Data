from bs4 import BeautifulSoup
from selenium import webdriver
import time, json, requests
from openpyxl import Workbook, load_workbook
import os

def getUID(htmlText):
    soup = BeautifulSoup(htmlText, 'html.parser')
    x = soup.find_all('td', {'class':'align-middle'})
    response = []
    c = 0
    for i in x:
        if c%5 ==2:
            # print(i.text)
            response.append(i.text)
        c += 1    
    return response

def GetCompanyId():
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_experimental_option("prefs", {"profile.password_manager_enabled": False, "credentials_enable_service": False})
    driver=webdriver.Chrome(executable_path=r'chromedriver.exe', chrome_options=options)
    driver.get("https://fbfh.trade.gov.tw/fb/web/queryBasicf.do")
    # 輸入資訊
    query = input('輸入"出進口貨品號列": ')
    checkCode = input('輸入驗證碼: ')
    
    queryInput = driver.find_element_by_id('q_ccc')
    queryInput.send_keys(query)

    checkCodeInput = driver.find_element_by_id('verifyCode')
    checkCodeInput.send_keys(checkCode)

    submitButton = driver.find_element_by_name('querySubmit')
    submitButton.click()

    time.sleep(3)

    companyList = []
    firstDo = True
    firstPage = 2
    while True:
        html = driver.page_source
        c = getUID(html)
        companyList.extend(c)
        try:
            nextPageButton = driver.find_element_by_xpath('//*[@id="listContainer"]/div/div[2]/div/div/div[2]/button[{0}]'.format(firstPage))
            nextPageButton.click()
            firstPage = 3
            time.sleep(3)
        except Exception as e:
            # print(e)
            break
    return companyList, query

def GetCompanyData(company_id):
    cc = []
    total_data = []
    
    for i in company_id:
        north_area = True
        c = {}
        data = ("{\"banNo\":\"%s\"}"%(i)).encode('utf-8')
        # basic data
        url = "https://fbfh.trade.gov.tw/fb/common/popBasic.action"
        headers = json.loads(r'''{
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua": "\"Chromium\";v=\"94\", \"Google Chrome\";v=\"94\", \";Not A Brand\";v=\"99\"",
            "x-requested-with": "XMLHttpRequest",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36",
            "referer": "https://fbfh.trade.gov.tw/fb/web/queryBasicf.do",
            "content-type": "application/json;charset=UTF-8",
            "accept": "application/json, text/javascript, */*; q=0.01"
        }''')
        resp = requests.post(url, data=data, headers=headers)
        x = resp.json()
        c['id'] = x['retrieveDataList'][0][0]
        c['name'] = x['retrieveDataList'][0][1]
        c['phone'] = x['retrieveDataList'][0][8]
        c['address'] = x['retrieveDataList'][0][6]
        # action data
        url = "https://fbfh.trade.gov.tw/fb/common/popGrade.action"
        headers = json.loads(r'''{
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua": "\"Chromium\";v=\"94\", \"Google Chrome\";v=\"94\", \";Not A Brand\";v=\"99\"",
            "x-requested-with": "XMLHttpRequest",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36",
            "referer": "https://fbfh.trade.gov.tw/fb/web/queryBasicf.do",
            "content-type": "application/json;charset=UTF-8",
            "accept": "application/json, text/javascript, */*; q=0.01"
        }''')
        resp1 = requests.post(url, data=data, headers=headers)
        y = resp1.json()
        rg = []
        for rd in y['retrieveDataList']:
            rg.append('{0}/{1}'.format(rd[4], rd[5]))
        c['range'] = rg
        # discard data which action is M/M this year
        city_list = ['臺北市', '新北市', '桃園市', '新竹縣', '新竹市', '宜蘭縣', '苗栗縣', '花蓮縣']
        if not c.get('phone', False):
            print('{0} {1} 未提供電話號碼，捨去資料'.format(c['id'], c['name']))
        else:
            if y['retrieveDataList'][0][4].upper() != 'M' or y['retrieveDataList'][0][5].upper() != 'M':
                if c['phone'] != '' and (c['phone'][:2] != '02' and c['phone'][:2] != '03'):
                    print('{0} {1} 電話開頭為 \'{2}\'，捨去資料'.format(c['id'], c['name'], c['phone'][:2]))
                elif c['address'] != '' and c['address'][:3] not in city_list:
                    print('{0} {1} 位在 \'{2}\'，捨去資料'.format(c['id'], c['name'], c['address'][:3]))
                elif c['phone'] == ''  and c['address'] == '':
                    print('{0} {1} 未提供電話及地址無法判斷，故加入列表'.format(c['id'], c['name']))
                    cc.append(c)
                else:
                    cc.append(c)
            else:
                print('{0} 年進出口實績為M/M，捨去資料---> [{1} {2}]'.format(y['retrieveDataList'][0][-1], y['retrieveDataList'][0][0], y['retrieveDataList'][0][2]))
        total_data.append(c)
    return cc, total_data

def CreateNewFile(item):
    wb = Workbook()
    # wb.create_sheet('sheet1', 0)
    sheet = wb['Sheet']
    sheet['A1'] = '統一編號'
    sheet['B1'] = '中文名稱'
    sheet['C1'] = '電話'
    sheet['D1'] = '實績1'
    sheet['E1'] = '實績2'
    sheet['F1'] = '實績3'
    sheet['G1'] = '實績4'
    sheet['H1'] = '實績5'
    sheet['I1'] = '地址'
    wb.save('./excel/{0}.xlsx'.format(item))

def ExportToExcel(cc, item, offset):
    wb = load_workbook('./excel/{0}.xlsx'.format(item))
    sheet = wb['Sheet']
    for i in range(len(cc)):
        col = i + offset + 1
        sheet[f'A{col}'] = cc[i].get('id', '')
        sheet[f'B{col}'] = cc[i].get('name', '')
        sheet[f'C{col}'] = cc[i].get('phone', '')
        sheet[f'I{col}'] = cc[i].get('address', '')
        try:
            sheet[f'D{col}'] = cc[i]['range'][0]
            sheet[f'E{col}'] = cc[i]['range'][1]
            sheet[f'F{col}'] = cc[i]['range'][2]
            sheet[f'G{col}'] = cc[i]['range'][3]
            sheet[f'H{col}'] = cc[i]['range'][4]
        except:
            pass        
    wb.save('./excel/{0}.xlsx'.format(item))


if __name__ == '__main__':
    pwd = os.getcwd()
    if not os.path.exists(os.path.join(pwd, 'log')):
        os.mkdir(os.path.join(pwd, 'log'))
    if not os.path.exists(os.path.join(pwd, 'excel')):
        os.mkdir(os.path.join(pwd, 'excel'))    
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_experimental_option("prefs", {"profile.password_manager_enabled": False, "credentials_enable_service": False})
    driver=webdriver.Chrome(executable_path=r'chromedriver.exe', chrome_options=options)
    driver.get("https://fbfh.trade.gov.tw/fb/web/queryBasicf.do")
    driver.execute_script("window.scrollTo(0, 1000)")
    # 輸入資訊
    query = input('輸入"出進口貨品號列": ')
    checkCode = input('輸入驗證碼: ')
    
    queryInput = driver.find_element_by_id('q_ccc')
    queryInput.send_keys(query)

    checkCodeInput = driver.find_element_by_id('verifyCode')
    checkCodeInput.send_keys(checkCode)

    submitButton = driver.find_element_by_name('querySubmit')
    submitButton.click()

    time.sleep(3)

    CreateNewFile(query)
    firstDo = True
    firstPage = 2
    thisPage = 1
    dataCount = 0
    log = []
    while True:
        print('...抓取第 {0} 頁資料'.format(thisPage))
        html = driver.page_source
        c = getUID(html)
        companyData, totalData = GetCompanyData(c)
        ExportToExcel(companyData, query, dataCount+1)
        log.append(totalData)
        dataCount += len(companyData)
        print('...第 {0} 頁 寫入完成 共 {1} 筆'.format(thisPage, dataCount))
        try:
            nextPageButton = driver.find_element_by_xpath('//*[@id="listContainer"]/div/div[2]/div/div/div[2]/button[{0}]'.format(firstPage))
            nextPageButton.click()
            firstPage = 3
            thisPage += 1
            time.sleep(3)
        except Exception as e:
            # print(e)
            break

    with open(f'./log/{query}.json', 'w+', encoding='utf-8') as f:
        f.write(json.dumps(log, ensure_ascii=False, indent=4))
    driver.quit()
    input('按任意鍵關閉視窗...')
